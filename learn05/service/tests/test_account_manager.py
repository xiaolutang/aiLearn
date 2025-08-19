#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试账号管理器
用于创建、管理和清理智能教学助手应用的测试账号
"""

import json
import hashlib
import sqlite3
import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional

class TestAccountManager:
    """测试账号管理器"""
    
    def __init__(self, db_path: str = None, accounts_file: str = None):
        """
        初始化测试账号管理器
        
        Args:
            db_path: 数据库文件路径
            accounts_file: 测试账号配置文件路径
        """
        self.db_path = db_path or "../student_database.db"
        self.accounts_file = accounts_file or "test_accounts.json"
        self.test_accounts = self._load_test_accounts()
        
    def _load_test_accounts(self) -> Dict:
        """加载测试账号配置"""
        try:
            with open(self.accounts_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"警告: 测试账号配置文件 {self.accounts_file} 不存在")
            return {"test_accounts": {"accounts": []}}
        except json.JSONDecodeError as e:
            print(f"错误: 解析测试账号配置文件失败: {e}")
            return {"test_accounts": {"accounts": []}}
    
    def _hash_password(self, password: str) -> str:
        """密码哈希处理"""
        return hashlib.sha256(password.encode()).hexdigest()
    
    def _get_db_connection(self) -> sqlite3.Connection:
        """获取数据库连接"""
        if not os.path.exists(self.db_path):
            raise FileNotFoundError(f"数据库文件不存在: {self.db_path}")
        return sqlite3.connect(self.db_path)
    
    def create_test_accounts(self) -> bool:
        """批量创建测试账号"""
        try:
            conn = self._get_db_connection()
            cursor = conn.cursor()
            
            # 创建用户表（如果不存在）
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id TEXT PRIMARY KEY,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    email TEXT UNIQUE NOT NULL,
                    role TEXT NOT NULL,
                    status TEXT DEFAULT 'active',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    last_login TIMESTAMP,
                    description TEXT
                )
            """)
            
            # 创建权限表（如果不存在）
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS user_permissions (
                    user_id TEXT,
                    permission TEXT,
                    granted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users (id)
                )
            """)
            
            accounts = self.test_accounts.get("test_accounts", {}).get("accounts", [])
            created_count = 0
            
            for account in accounts:
                try:
                    # 插入用户基本信息
                    cursor.execute("""
                        INSERT OR REPLACE INTO users 
                        (id, username, password_hash, email, role, status, description)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (
                        account["id"],
                        account["username"],
                        self._hash_password(account["password"]),
                        account["email"],
                        account["role"],
                        account["status"],
                        account["description"]
                    ))
                    
                    # 删除旧的权限记录
                    cursor.execute("DELETE FROM user_permissions WHERE user_id = ?", (account["id"],))
                    
                    # 插入权限信息
                    for permission in account["permissions"]:
                        cursor.execute("""
                            INSERT INTO user_permissions (user_id, permission)
                            VALUES (?, ?)
                        """, (account["id"], permission))
                    
                    created_count += 1
                    print(f"✅ 创建测试账号: {account['username']} ({account['role']})")
                    
                except sqlite3.IntegrityError as e:
                    print(f"⚠️  账号 {account['username']} 创建失败: {e}")
                    continue
            
            conn.commit()
            conn.close()
            
            print(f"\n🎉 成功创建 {created_count} 个测试账号")
            return True
            
        except Exception as e:
            print(f"❌ 创建测试账号失败: {e}")
            return False
    
    def verify_test_accounts(self) -> Dict[str, bool]:
        """验证测试账号是否正确创建"""
        try:
            conn = self._get_db_connection()
            cursor = conn.cursor()
            
            accounts = self.test_accounts.get("test_accounts", {}).get("accounts", [])
            verification_results = {}
            
            for account in accounts:
                # 检查用户是否存在
                cursor.execute(
                    "SELECT id, username, role, status FROM users WHERE username = ?",
                    (account["username"],)
                )
                user_result = cursor.fetchone()
                
                if user_result:
                    # 检查权限是否正确
                    cursor.execute(
                        "SELECT permission FROM user_permissions WHERE user_id = ?",
                        (user_result[0],)
                    )
                    permissions = [row[0] for row in cursor.fetchall()]
                    
                    expected_permissions = set(account["permissions"])
                    actual_permissions = set(permissions)
                    
                    verification_results[account["username"]] = {
                        "exists": True,
                        "role_correct": user_result[2] == account["role"],
                        "status_correct": user_result[3] == account["status"],
                        "permissions_correct": expected_permissions == actual_permissions,
                        "missing_permissions": list(expected_permissions - actual_permissions),
                        "extra_permissions": list(actual_permissions - expected_permissions)
                    }
                else:
                    verification_results[account["username"]] = {
                        "exists": False,
                        "role_correct": False,
                        "status_correct": False,
                        "permissions_correct": False
                    }
            
            conn.close()
            return verification_results
            
        except Exception as e:
            print(f"❌ 验证测试账号失败: {e}")
            return {}
    
    def cleanup_test_accounts(self) -> bool:
        """清理测试账号"""
        try:
            conn = self._get_db_connection()
            cursor = conn.cursor()
            
            accounts = self.test_accounts.get("test_accounts", {}).get("accounts", [])
            deleted_count = 0
            
            for account in accounts:
                # 删除权限记录
                cursor.execute("DELETE FROM user_permissions WHERE user_id = ?", (account["id"],))
                
                # 删除用户记录
                cursor.execute("DELETE FROM users WHERE id = ?", (account["id"],))
                
                deleted_count += 1
                print(f"🗑️  删除测试账号: {account['username']}")
            
            conn.commit()
            conn.close()
            
            print(f"\n🧹 成功清理 {deleted_count} 个测试账号")
            return True
            
        except Exception as e:
            print(f"❌ 清理测试账号失败: {e}")
            return False
    
    def get_accounts_by_role(self, role: str) -> List[Dict]:
        """根据角色获取测试账号"""
        accounts = self.test_accounts.get("test_accounts", {}).get("accounts", [])
        return [acc for acc in accounts if acc["role"] == role]
    
    def get_accounts_by_scenario(self, scenario: str) -> List[Dict]:
        """根据测试场景获取测试账号"""
        scenarios = self.test_accounts.get("test_accounts", {}).get("test_scenarios", {})
        if scenario in scenarios:
            account_ids = scenarios[scenario].get("test_accounts", [])
            accounts = self.test_accounts.get("test_accounts", {}).get("accounts", [])
            return [acc for acc in accounts if acc["id"] in account_ids]
        return []
    
    def print_account_summary(self):
        """打印账号摘要信息"""
        accounts = self.test_accounts.get("test_accounts", {}).get("accounts", [])
        
        print("\n📊 测试账号摘要")
        print("=" * 50)
        
        # 按角色统计
        role_counts = {}
        status_counts = {}
        
        for account in accounts:
            role = account["role"]
            status = account["status"]
            
            role_counts[role] = role_counts.get(role, 0) + 1
            status_counts[status] = status_counts.get(status, 0) + 1
        
        print("\n👥 按角色分布:")
        for role, count in role_counts.items():
            print(f"  {role}: {count} 个")
        
        print("\n📈 按状态分布:")
        for status, count in status_counts.items():
            print(f"  {status}: {count} 个")
        
        print(f"\n📝 总计: {len(accounts)} 个测试账号")
        
        # 测试场景
        scenarios = self.test_accounts.get("test_accounts", {}).get("test_scenarios", {})
        print(f"\n🎯 测试场景: {len(scenarios)} 个")
        for scenario_name in scenarios.keys():
            print(f"  - {scenario_name}")
    
    def export_login_credentials(self, output_file: str = "test_login_credentials.txt"):
        """导出登录凭据到文件"""
        try:
            accounts = self.test_accounts.get("test_accounts", {}).get("accounts", [])
            
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write("智能教学助手 - 测试账号登录凭据\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                
                for account in accounts:
                    f.write(f"角色: {account['role']}\n")
                    f.write(f"用户名: {account['username']}\n")
                    f.write(f"密码: {account['password']}\n")
                    f.write(f"邮箱: {account['email']}\n")
                    f.write(f"状态: {account['status']}\n")
                    f.write(f"描述: {account['description']}\n")
                    f.write("-" * 30 + "\n\n")
            
            print(f"✅ 登录凭据已导出到: {output_file}")
            
        except Exception as e:
            print(f"❌ 导出登录凭据失败: {e}")
    
    def export_to_excel(self, output_file: str = "test_accounts.xlsx") -> bool:
        """导出测试账号到Excel表格"""
        try:
            accounts = self.test_accounts.get("test_accounts", {}).get("accounts", [])
            
            # 准备数据
            data = []
            for i, account in enumerate(accounts, 1):
                row = {
                    '序号': i,
                    '角色': account['role'],
                    '用户名': account['username'],
                    '密码': account['password'],
                    '邮箱': account['email'],
                    '状态': account['status'],
                    '权限': ', '.join(account['permissions']),
                    '描述': account['description']
                }
                data.append(row)
            
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 创建Excel写入器
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 写入主表
                df.to_excel(writer, sheet_name='测试账号', index=False)
                
                # 创建角色统计表
                role_stats = df['角色'].value_counts().reset_index()
                role_stats.columns = ['角色', '数量']
                role_stats.to_excel(writer, sheet_name='角色统计', index=False)
                
                # 创建状态统计表
                status_stats = df['状态'].value_counts().reset_index()
                status_stats.columns = ['状态', '数量']
                status_stats.to_excel(writer, sheet_name='状态统计', index=False)
                
                # 获取工作表并设置格式
                workbook = writer.book
                worksheet = writer.sheets['测试账号']
                
                # 设置列宽
                column_widths = {
                    'A': 8,   # 序号
                    'B': 15,  # 角色
                    'C': 20,  # 用户名
                    'D': 15,  # 密码
                    'E': 25,  # 邮箱
                    'F': 10,  # 状态
                    'G': 30,  # 权限
                    'H': 40   # 描述
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 设置表头样式
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 设置数据行样式
                data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                for row in worksheet.iter_rows(min_row=2, max_row=len(data)+1):
                    for cell in row:
                        cell.alignment = data_alignment
            
            print(f"✅ 测试账号已导出到Excel: {output_file}")
            print(f"包含 {len(data)} 个测试账号，分为3个工作表：测试账号、角色统计、状态统计")
            return True
            
        except Exception as e:
            print(f"❌ 导出Excel失败: {e}")
            print("请确保已安装pandas和openpyxl: pip install pandas openpyxl")
            return False

def main():
    """主函数 - 命令行界面"""
    import argparse
    
    parser = argparse.ArgumentParser(description="智能教学助手测试账号管理器")
    parser.add_argument("action", choices=["create", "verify", "cleanup", "summary", "export"],
                       help="操作类型")
    parser.add_argument("--db", default="../student_database.db", help="数据库文件路径")
    parser.add_argument("--accounts", default="test_accounts.json", help="测试账号配置文件")
    parser.add_argument("--output", default="test_login_credentials.txt", help="导出文件路径 (.txt或.xlsx格式)")
    
    args = parser.parse_args()
    
    manager = TestAccountManager(args.db, args.accounts)
    
    if args.action == "create":
        print("🚀 开始创建测试账号...")
        success = manager.create_test_accounts()
        if success:
            print("\n✅ 测试账号创建完成")
        else:
            print("\n❌ 测试账号创建失败")
    
    elif args.action == "verify":
        print("🔍 开始验证测试账号...")
        results = manager.verify_test_accounts()
        
        print("\n📋 验证结果:")
        for username, result in results.items():
            status = "✅" if all([result["exists"], result["role_correct"], 
                               result["status_correct"], result["permissions_correct"]]) else "❌"
            print(f"{status} {username}: 存在={result['exists']}, 角色正确={result['role_correct']}, "
                  f"状态正确={result['status_correct']}, 权限正确={result['permissions_correct']}")
    
    elif args.action == "cleanup":
        print("🧹 开始清理测试账号...")
        confirm = input("确认要删除所有测试账号吗？(y/N): ")
        if confirm.lower() == 'y':
            success = manager.cleanup_test_accounts()
            if success:
                print("\n✅ 测试账号清理完成")
            else:
                print("\n❌ 测试账号清理失败")
        else:
            print("❌ 操作已取消")
    
    elif args.action == "summary":
        manager.print_account_summary()
    
    elif args.action == "export":
        print("📤 导出登录凭据...")
        if args.output.endswith('.xlsx'):
            manager.export_to_excel(args.output)
        else:
            manager.export_login_credentials(args.output)

if __name__ == "__main__":
    main()